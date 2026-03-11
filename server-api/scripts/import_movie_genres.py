#!/usr/bin/env python3
<<<<<<< codex/enhance-recommendation-algorithm-with-item-based-filtering-n7s1e8
"""将 id.xlsx 中的 genres_cn 写入 movies.genres 字段。

Excel 列格式示例：
id | imdb_id | genres_cn
16612 | tt0000001 | 纪录
88013 | tt0000003 | 喜剧, 动画

说明：
- 仅写入类型名称，不写入类型 id。
- genres_cn 会被拆分并存成 JSON 数组，例如：["喜剧", "动画"]。
- 通过 id 或 imdb_id 匹配 movies 表。
"""

import argparse
import json
from typing import List

import pymysql

try:
  from openpyxl import load_workbook
except ImportError as exc:
  raise SystemExit(
    '缺少依赖 openpyxl，请先安装：pip install openpyxl'
  ) from exc
=======
"""将 CSV 中的电影类型写入 movies.genres 字段。

CSV 示例：
genres,id,imdb_id
"[{'id': 12, 'name': 'Adventure'}, {'id': 16, 'name': 'Animation'}]",2,tt0094675
"""

import argparse
import ast
import csv
import json

import pymysql


def parse_genre_names(raw_genres: str):
  if not raw_genres:
    return []

  parsed = ast.literal_eval(raw_genres)
  if not isinstance(parsed, list):
    return []

  names = []
  for item in parsed:
    if isinstance(item, dict) and item.get('name'):
      names.append(str(item['name']).strip())

  return names
>>>>>>> main


def ensure_genres_column(conn, database_name: str):
  with conn.cursor() as cursor:
    cursor.execute(
      """
      SELECT COUNT(*)
      FROM information_schema.COLUMNS
      WHERE TABLE_SCHEMA = %s
        AND TABLE_NAME = 'movies'
        AND COLUMN_NAME = 'genres'
      """,
      (database_name,)
    )
    exists = cursor.fetchone()[0] > 0

    if not exists:
      cursor.execute('ALTER TABLE movies ADD COLUMN genres JSON NULL')
      conn.commit()
      print('已新增 movies.genres 字段')


<<<<<<< codex/enhance-recommendation-algorithm-with-item-based-filtering-n7s1e8
def parse_genres_cn(genres_cn: str) -> List[str]:
  if genres_cn is None:
    return []

  text = str(genres_cn).strip()
  if not text:
    return []

  # 兼容中文/英文逗号、顿号、分号等分隔符
  separators = [',', '，', '、', ';', '；', '|', '/']
  normalized = text
  for sep in separators:
    normalized = normalized.replace(sep, ',')

  genres = [item.strip() for item in normalized.split(',') if item.strip()]

  # 去重并保持顺序
  seen = set()
  unique_genres = []
  for genre in genres:
    if genre not in seen:
      seen.add(genre)
      unique_genres.append(genre)

  return unique_genres


def load_rows_from_xlsx(xlsx_path: str):
  wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
  ws = wb.active

  rows_iter = ws.iter_rows(values_only=True)
  header = next(rows_iter, None)
  if not header:
    raise ValueError('Excel 文件为空，未找到表头')

  header_map = {str(col).strip(): idx for idx, col in enumerate(header) if col is not None}
  required_columns = ['id', 'imdb_id', 'genres_cn']
  for col in required_columns:
    if col not in header_map:
      raise ValueError(f'缺少必需列: {col}')

  parsed_rows = []
  for raw in rows_iter:
    movie_id = raw[header_map['id']] if header_map['id'] < len(raw) else None
    imdb_id = raw[header_map['imdb_id']] if header_map['imdb_id'] < len(raw) else None
    genres_cn = raw[header_map['genres_cn']] if header_map['genres_cn'] < len(raw) else None

    if movie_id is None and imdb_id is None:
      continue

    parsed_rows.append(
      {
        'id': str(movie_id).strip() if movie_id is not None else '',
        'imdb_id': str(imdb_id).strip() if imdb_id is not None else '',
        'genres_cn': parse_genres_cn(genres_cn)
      }
    )

  wb.close()
  return parsed_rows


def import_genres_from_xlsx(conn, xlsx_path: str):
  rows = load_rows_from_xlsx(xlsx_path)
  updated = 0

  with conn.cursor() as cursor:
    for row in rows:
      genres_json = json.dumps(row['genres_cn'], ensure_ascii=False)
      cursor.execute(
        """
        UPDATE movies
        SET genres = %s
        WHERE id = %s OR imdb_id = %s
        """,
        (genres_json, row['id'], row['imdb_id'])
      )
      updated += cursor.rowcount
=======
def import_genres(conn, csv_path: str):
  updated = 0

  with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
    reader = csv.DictReader(f)

    required_cols = {'genres', 'id', 'imdb_id'}
    if not required_cols.issubset(set(reader.fieldnames or [])):
      raise ValueError('CSV 必须包含 genres, id, imdb_id 三列')

    with conn.cursor() as cursor:
      for row in reader:
        movie_id = row.get('id')
        imdb_id = row.get('imdb_id')
        genre_names = parse_genre_names(row.get('genres', ''))

        if not movie_id and not imdb_id:
          continue

        genres_json = json.dumps(genre_names, ensure_ascii=False)

        cursor.execute(
          """
          UPDATE movies
          SET genres = %s
          WHERE id = %s OR imdb_id = %s
          """,
          (genres_json, movie_id, imdb_id)
        )
        updated += cursor.rowcount
>>>>>>> main

  conn.commit()
  return updated


def main():
<<<<<<< codex/enhance-recommendation-algorithm-with-item-based-filtering-n7s1e8
  parser = argparse.ArgumentParser(description='将 id.xlsx 中的 genres_cn 导入 movies.genres')
  parser.add_argument('--xlsx', required=True, help='Excel 文件路径，例如 ./id.xlsx')
=======
  parser = argparse.ArgumentParser(description='导入电影类型到 MySQL movies 表')
  parser.add_argument('--csv', required=True, help='CSV 文件路径')
>>>>>>> main
  parser.add_argument('--host', default='localhost')
  parser.add_argument('--port', type=int, default=3306)
  parser.add_argument('--user', default='root')
  parser.add_argument('--password', default='123123lzy')
  parser.add_argument('--database', default='movie_db')
  args = parser.parse_args()

  conn = pymysql.connect(
    host=args.host,
    port=args.port,
    user=args.user,
    password=args.password,
    database=args.database,
    charset='utf8mb4'
  )

  try:
    ensure_genres_column(conn, args.database)
<<<<<<< codex/enhance-recommendation-algorithm-with-item-based-filtering-n7s1e8
    updated_count = import_genres_from_xlsx(conn, args.xlsx)
=======
    updated_count = import_genres(conn, args.csv)
>>>>>>> main
    print(f'导入完成，更新行数: {updated_count}')
  finally:
    conn.close()


if __name__ == '__main__':
  main()
